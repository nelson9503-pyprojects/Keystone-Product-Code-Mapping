import os
import json
import openpyxl


class MapController:

    def __init__(self):
        self.open_map_json()

    def check_map_json_exist(self):
        if not os.path.exists("product_code_map.json"):
            j = {"standard": [], "mapping": {}}
            with open("product_code_map.json", 'w') as f:
                f.write(json.dumps(j))

    def open_map_json(self):
        self.check_map_json_exist()
        with open("product_code_map.json", 'r') as f:
            j = json.loads(f.read())
            self.standard = j["standard"]
            self.map = j["mapping"]

    def save_map_json(self):
        j = {
            "standard": self.standard,
            "mapping": self.map
        }
        with open("product_code_map.json", 'w') as f:
            f.write(json.dumps(j))

    def query(self, code: str) -> any:
        """
        if code is standard code, return True.
        if code is non standard code, return the mapped standard code.
        if code is not in record, return False.
        """
        if code in self.standard:
            return True
        elif code in self.map:
            return self.map[code]
        else:
            return False

    def add_standard(self, code: str) -> bool:
        """
        if code is non-standard code, it will not be added to standard list and return False.
        """
        if code == None or code == "":
            return False
        if code in self.map:
            return False
        if not code in self.standard:
            self.standard.append(code)
        self.save_map_json()
        return True

    def add_non_standard(self, code: str, mapcode: str) -> bool:
        """
        if code is standard code, it will not be added to non-standard map and return False.
        """
        if code == None or code == "":
            return False
        if code in self.standard:
            return False
        if mapcode == None or mapcode == "":
            return False
        if not mapcode in self.standard:
            return False
        self.map[code] = mapcode
        self.save_map_json()
        return True

    def export_excel(self, path: str) -> bool:
        try:
            # mapping
            wb = openpyxl.Workbook()
            sh = wb.create_sheet("Mapping", 0)
            sh.cell(1, 1).value = "Non Standard Codes"
            sh.cell(1, 2).value = "Standard Codes"
            row = 2
            for code in self.map:
                sh.cell(row, 1).value = code
                sh.cell(row, 2).value = self.map[code]
                row += 1
            # standard
            sh = wb.create_sheet("Standard_Codes", 1)
            sh.cell(1, 1).value = "Registered Standard Codes"
            row = 2
            for code in self.standard:
                sh.cell(row, 1).value = code
                row += 1
            wb.save(path)
        except:
            return False
        return True

    def import_from_excel(self, path: str) -> any:
        cache_standard = []
        cache_map = {}
        try:
            wb = openpyxl.open(path)
            sh = wb["Mapping"]
            row = 2
            while not sh.cell(row, 1).value == None:
                code = sh.cell(row, 2).value
                mapcode = sh.cell(row, 1).value
                if code in cache_map:
                    return "Code assiged to non-standard list cannot become standard code"
                if mapcode in cache_standard:
                    return "Code assiged to stanard list cannot become non-standard code"
                if mapcode == None:
                    return "Error: empty cell is found"
                if not code in cache_standard:
                    cache_standard.append(code)
                cache_map[mapcode] = code
                row += 1
            sh = wb["Standard_Codes"]
            row = 2
            while not sh.cell(row, 1).value == None:
                code = sh.cell(row, 1).value
                if not code in cache_standard:
                    cache_standard.append(code)
                row += 1
        except:
            return False
        self.standard = cache_standard
        self.map = cache_map
        self.save_map_json()
        return True

    def batch_mapping(self, path: str):
        try:
            wb = openpyxl.open(path)
            sh = wb[wb.sheetnames[0]]
            row = 1
            while not sh.cell(row, 1).value == None:
                code = sh.cell(row, 1).value
                code = code.upper()
                result = self.query(code)
                if result == True:
                    sh.cell(row, 2).value = code
                elif not result == False:
                    sh.cell(row, 2).value = result
                row += 1
            wb.save(path)
        except:
            return False
        return True
